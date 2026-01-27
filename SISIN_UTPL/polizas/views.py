from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse, FileResponse
from django.forms import inlineformset_factory
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.contrib import messages
from .models import Poliza, RamoPoliza, Deducible, BienAsegurado, Zona
from .forms import RamoPolizaForm, PolizaForm, DeducibleForm
from usuarios.models import Usuario
from django.db import transaction
import os
import logging
import csv
import json
from datetime import datetime, timedelta
from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Font
from django.db import transaction

logger = logging.getLogger(__name__)

# Create your views here.

def ver_polizas(request):
    polizas = Poliza.objects.select_related('titular', 'bien').prefetch_related('ramos', 'deducibles').all()
    
    # Calculate statistics
    total_polizas = polizas.count()
    polizas_activas = polizas.filter(estado='activa').count()
    polizas_por_vencer = polizas.filter(estado='suspendida').count()
    prima_total = sum(poliza.prima for poliza in polizas)
    
    context = {
        'polizas': polizas,
        'total_polizas': total_polizas,
        'polizas_activas': polizas_activas,
        'polizas_por_vencer': polizas_por_vencer,
        'prima_total': prima_total
    }
    
    return render(request, 'asesora/polizas.html', context)

def descargar_pdf(request, poliza_id):
    poliza = get_object_or_404(Poliza, id=poliza_id)
    
    if poliza.pdf_file:
        # Si existe un PDF, descargarlo
        response = FileResponse(poliza.pdf_file, as_attachment=True)
        response['Content-Disposition'] = f'attachment; filename="{poliza.numero_poliza}.pdf"'
        return response
    else:
        # Si no existe PDF, generar uno simple
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import letter
        from io import BytesIO
        import tempfile
        
        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=letter)
        
        # T铆tulo
        p.setFont("Helvetica-Bold", 16)
        p.drawString(100, 750, f"PLIZA DE SEGURO")
        
        # Informaci贸n de la p贸liza
        p.setFont("Helvetica", 12)
        p.drawString(100, 700, f"N煤mero de P贸liza: {poliza.numero_poliza}")
        p.drawString(100, 680, f"Tipo de P贸liza: {poliza.tipo_poliza}")
        p.drawString(100, 660, f"Aseguradora: {poliza.aseguradora}")
        p.drawString(100, 640, f"Prima: ${poliza.prima}")
        p.drawString(100, 620, f"Fecha de Inicio: {poliza.fecha_inicio}")
        p.drawString(100, 600, f"Fecha de Fin: {poliza.fecha_fin}")
        
        # Deducibles
        p.drawString(100, 560, "DEDUCIBLES:")
        y_pos = 540
        for deducible in poliza.deducibles.all():
            p.drawString(120, y_pos, f"- {deducible.concepto}: {deducible.porcentaje}%" if deducible.porcentaje else f"- {deducible.concepto}: ${deducible.monto}")
            y_pos -= 20
        
        p.save()
        
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{poliza.numero_poliza}.pdf"'
        return response

@csrf_exempt
@require_http_methods(["POST"])
def crear_poliza(request):
    try:
        with transaction.atomic():  # Si algo falla, no se guarda nada
            data = request.POST
            files = request.FILES
            
            logger.info(f"Datos recibidos: {dict(data)}")
            
            # 1. Obtener o crear la Zona
            zona_nombre = data.get('zona', 'Sin especificar')
            zona, _ = Zona.objects.get_or_create(nombre=zona_nombre)
            
            # 2. Mapear tipo de bien a entero
            tipo_bien_map = {
                'Residencial': 1,
                'Comercial': 2,
                'Industrial': 3,
            }
            tipo_bien_str = data.get('tipo_bien', 'Residencial')
            tipo_bien_int = tipo_bien_map.get(tipo_bien_str, 1)
            
            # 3. Crear el Bien Asegurado
            bien = BienAsegurado.objects.create(
                descripcion=data.get('descripcion_bien', ''),
                estado=1,  # Estado activo por defecto
                tipo_bien=tipo_bien_int,
                zona=zona,
                valor=data.get('valor_bien', 0) or 0
            )

            # 4. Obtener o crear un usuario titular
            # Primero intentamos obtener el primer usuario existente
            titular = Usuario.objects.first()
            if not titular:
                # Si no existe ning煤n usuario, creamos uno por defecto
                titular, created = Usuario.objects.get_or_create(
                    email='usuario@default.com',
                    defaults={
                        'nombre': 'Usuario',
                        'apellido': 'Por Defecto',
                        'telefono': ''
                    }
                )
                if created:
                    logger.info(f"Usuario por defecto creado con id: {titular.id}")

            # 5. Crear la P贸liza
            poliza = Poliza.objects.create(
                numero_poliza=data.get('numero_poliza'),
                aseguradora=data.get('aseguradora', ''),
                tipo_poliza=tipo_bien_str,  # Usar el tipo de bien como tipo de p贸liza
                fecha_emision=data.get('fecha_emision'),
                fecha_vencimiento=data.get('fecha_fin'),  # fecha_vencimiento = fecha_fin
                fecha_inicio=data.get('fecha_inicio'),
                fecha_fin=data.get('fecha_fin'),
                prima=data.get('total_prima', 0) or 0,
                cobertura=data.get('descripcion_bien', ''),  # Usar descripci贸n como cobertura
                bien=bien,
                titular=titular,  # Usar el usuario obtenido o creado
                pdf_file=files.get('pdf_file')
            )

            # 6. Procesar Ramos (JSON enviado desde JS)
            ramos_data = json.loads(data.get('ramos_json', '[]'))
            for item in ramos_data:
                RamoPoliza.objects.create(
                    poliza=poliza,
                    grupo=item.get('grupo', ''),
                    subgrupo=item.get('subgrupo', ''),
                    ramo=item.get('ramo', ''),
                    suma_asegurada=item.get('suma_asegurada', 0) or 0,
                    prima=0,  # Valor por defecto
                    base_imponible=0,
                    iva=0,
                    total_facturado=0,
                    deducible_minimo=0,
                    deducible_porcentaje=item.get('deducible_porcentaje', 0) or 0
                )

            # 7. Procesar Deducibles (JSON enviado desde JS)
            deducibles_data = json.loads(data.get('deducibles_json', '[]'))
            for item in deducibles_data:
                # Intentar convertir monto a decimal, si falla usar 0
                monto_str = item.get('monto', '0')
                try:
                    # Remover caracteres no num茅ricos excepto punto y coma
                    monto_clean = ''.join(c for c in str(monto_str) if c.isdigit() or c in '.,')
                    monto_clean = monto_clean.replace(',', '.')
                    monto = float(monto_clean) if monto_clean else 0
                except (ValueError, TypeError):
                    monto = 0
                    
                Deducible.objects.create(
                    poliza=poliza,
                    concepto=item.get('concepto', ''),
                    monto=monto
                )

            logger.info(f"P贸liza {poliza.numero_poliza} creada exitosamente")
            # Agregar mensaje de 茅xito para mostrar despu茅s de la redirecci贸n
            messages.success(request, f'隆P贸liza {poliza.numero_poliza} creada exitosamente!')
            return JsonResponse({'success': True, 'message': 'P贸liza y Bien creados con 茅xito', 'poliza_id': poliza.id})

    except Exception as e:
        logger.error(f"Error al crear p贸liza: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        messages.error(request, f'Error al crear la p贸liza: {str(e)}')
        return JsonResponse({'success': False, 'message': str(e)})

@csrf_exempt
@require_http_methods(["DELETE"])
def eliminar_poliza(request, poliza_id):
    try:
        poliza = get_object_or_404(Poliza, id=poliza_id)
        logger.info(f"Intentando eliminar p贸liza {poliza_id}: {poliza.numero_poliza}")
        
        # Eliminar deducibles relacionados primero
        poliza.deducibles.all().delete()
        
        # Eliminar la p贸liza
        poliza.delete()
        
        logger.info(f"P贸liza {poliza_id} eliminada exitosamente")
        return JsonResponse({'success': True, 'message': 'P贸liza eliminada exitosamente'})
    except Poliza.DoesNotExist:
        logger.error(f"La p贸liza {poliza_id} no existe")
        return JsonResponse({'success': False, 'message': 'La p贸liza no existe'})
    except Exception as e:
        logger.error(f"Error al eliminar p贸liza {poliza_id}: {str(e)}")
        return JsonResponse({'success': False, 'message': f'Error al eliminar la p贸liza: {str(e)}'})

@csrf_exempt
@require_http_methods(["POST"])
def editar_poliza(request, poliza_id):
    poliza = get_object_or_404(Poliza, id=poliza_id)

    try:
        with transaction.atomic():
            data = request.POST
            pdf_file = request.FILES.get('pdf_file')

            # Map tipo_bien string to integer
            tipo_bien_map = {
                'Residencial': 1,
                'Comercial': 2,
                'Industrial': 3,
            }
            tipo_bien_str = data.get('tipo_bien', 'Residencial')
            tipo_bien_int = tipo_bien_map.get(tipo_bien_str, 1)

            # Update Bien Asegurado
            bien = poliza.bien
            if bien:
                bien.descripcion = data.get('descripcion_bien', bien.descripcion)
                bien.tipo_bien = tipo_bien_int
                bien.valor = data.get('valor_bien', bien.valor) or 0

                # Update zona if changed
                zona_nombre = data.get('zona', '')
                if zona_nombre:
                    zona, _ = Zona.objects.get_or_create(nombre=zona_nombre)
                    bien.zona = zona

                bien.save()

            # Update P贸liza
            poliza.numero_poliza = data.get('numero_poliza', poliza.numero_poliza)
            poliza.aseguradora = data.get('aseguradora', poliza.aseguradora)
            poliza.tipo_poliza = tipo_bien_str
            poliza.fecha_emision = data.get('fecha_emision', poliza.fecha_emision)
            poliza.fecha_inicio = data.get('fecha_inicio', poliza.fecha_inicio)
            poliza.fecha_fin = data.get('fecha_fin', poliza.fecha_fin)
            poliza.prima = data.get('total_prima', poliza.prima) or 0
            poliza.cobertura = data.get('descripcion_bien', poliza.cobertura)

            # Update PDF if uploaded
            if pdf_file:
                poliza.pdf_file = pdf_file

            poliza.save()

            # Update Ramos
            poliza.ramos.all().delete()  # Delete existing ramos
            ramos_data = json.loads(data.get('ramos_json', '[]'))
            for item in ramos_data:
                RamoPoliza.objects.create(
                    poliza=poliza,
                    grupo=item.get('grupo', ''),
                    subgrupo=item.get('subgrupo', ''),
                    ramo=item.get('ramo', ''),
                    suma_asegurada=item.get('suma_asegurada', 0) or 0,
                    prima=0,  # Default value
                    base_imponible=0,
                    iva=0,
                    total_facturado=0,
                    deducible_minimo=0,
                    deducible_porcentaje=item.get('deducible_porcentaje', 0) or 0
                )

            # Update Deducibles
            poliza.deducibles.all().delete()
            deducibles_data = json.loads(data.get('deducibles_json', '[]'))
            for item in deducibles_data:
                monto_str = item.get('monto', '0')
                try:
                    monto_clean = ''.join(c for c in str(monto_str) if c.isdigit() or c in '.,')
                    monto_clean = monto_clean.replace(',', '.')
                    monto = float(monto_clean) if monto_clean else 0
                except (ValueError, TypeError):
                    monto = 0

                Deducible.objects.create(
                    poliza=poliza,
                    concepto=item.get('concepto', ''),
                    monto=monto
                )

            logger.info(f"P贸liza {poliza.numero_poliza} actualizada exitosamente")
            return JsonResponse({'success': True, 'message': 'P贸liza actualizada exitosamente'})

    except Exception as e:
        logger.error(f"Error al actualizar p贸liza {poliza_id}: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return JsonResponse({'success': False, 'message': str(e)})

def obtener_poliza(request, poliza_id):
    poliza = get_object_or_404(Poliza, id=poliza_id)

    bien = poliza.bien
    zona = bien.zona if bien and bien.zona else None

    # Map tipo_bien integer to string for the form
    tipo_bien_map = {
        1: 'Residencial',
        2: 'Comercial',
        3: 'Industrial',
    }
    tipo_bien_str = tipo_bien_map.get(bien.tipo_bien, 'Residencial') if bien else 'Residencial'

    data = {
        'id': poliza.id,
        'numero_poliza': poliza.numero_poliza,
        'tipo_poliza': poliza.tipo_poliza,
        'aseguradora': poliza.aseguradora,
        'cobertura': poliza.cobertura,
        'fecha_emision': poliza.fecha_emision.strftime('%Y-%m-%d') if poliza.fecha_emision else '',
        'fecha_inicio': poliza.fecha_inicio.strftime('%Y-%m-%d') if poliza.fecha_inicio else '',
        'fecha_fin': poliza.fecha_fin.strftime('%Y-%m-%d') if poliza.fecha_fin else '',
        'total_prima': str(poliza.prima),

        # Flatten bien data to match form expectations
        'descripcion_bien': bien.descripcion if bien else '',
        'valor_bien': str(bien.valor) if bien else '',
        'tipo_bien': tipo_bien_str,
        'zona': zona.nombre if zona else '',

        #  RAMOS (PUEDE ESTAR VACO)
        'ramos': [
            {
                'grupo': r.grupo,
                'subgrupo': r.subgrupo,
                'ramo': r.ramo,
                'suma_asegurada': str(r.suma_asegurada),
                'deducible_porcentaje': str(r.deducible_porcentaje)
            } for r in poliza.ramos.all()
        ],

        #  DEDUCIBLES (PUEDE ESTAR VACO)
        'deducibles': [
            {
                'concepto': d.concepto,
                'monto': str(d.monto)
            } for d in poliza.deducibles.all()
        ]
    }

    return JsonResponse(data)


def formulario_crear_poliza(request):
    """Vista para mostrar el formulario de crear p贸liza"""
    return render(request, 'asesora/crear_poliza.html')


def crear_poliza_old(request):
    # 1. Definimos el formset
    RamoFormSet = inlineformset_factory(
        Poliza, 
        RamoPoliza, 
        form=RamoPolizaForm,
        extra=1, 
        can_delete=True
    )
    
    if request.method == 'POST':
        # Aqu铆 recibimos los datos
        form = PolizaForm(request.POST)
        formset = RamoFormSet(request.POST)
        
        # Por ahora, como solo quieres que redirija sin guardar:
        return redirect('nombre_de_tu_url_de_exito') 
        
    else:
        # Carga inicial (GET)
        form = PolizaForm()
        formset = RamoFormSet()
    
    context = {
        'form': form,
        'formset': formset,
        'clientes': [], # O Cliente.objects.all()
    }
    
    return render(request, 'asesora/crear_poliza.html', context)

@csrf_exempt
@require_http_methods(["POST"])
def exportar_excel(request):
    try:
        polizas = Poliza.objects.select_related('titular', 'bien').prefetch_related('deducibles').all()
        
        # Crear archivo Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "P贸lizas"
        
        # Estilos
        header_font = Font(bold=True)
        
        # Encabezados
        headers = ['N潞 P贸liza', 'Tipo P贸liza', 'Titular', 'Fecha Inicio', 'Fecha Fin', 'Prima', 'Estado', 'Aseguradora']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
        
        # Datos
        for row, poliza in enumerate(polizas, 2):
            ws.cell(row=row, column=1, value=poliza.numero_poliza)
            ws.cell(row=row, column=2, value=poliza.tipo_poliza)
            ws.cell(row=row, column=3, value=str(poliza.titular))
            ws.cell(row=row, column=4, value=poliza.fecha_inicio.strftime('%d/%m/%Y') if poliza.fecha_inicio else '')
            ws.cell(row=row, column=5, value=poliza.fecha_fin.strftime('%d/%m/%Y') if poliza.fecha_fin else '')
            ws.cell(row=row, column=6, value=float(poliza.prima))
            ws.cell(row=row, column=7, value=poliza.get_estado_display())
            ws.cell(row=row, column=8, value=poliza.aseguradora)
        
        # Guardar en memoria
        from io import BytesIO
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Respuesta
        response = HttpResponse(
            excel_buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="polizas_export.xlsx"'
        
        return response
        
    except Exception as e:
        logger.error(f"Error exportando Excel: {e}")
        return JsonResponse({'success': False, 'message': f'Error al exportar: {str(e)}'})

@csrf_exempt
@require_http_methods(["POST"])
def renovar_vencidas(request):
    try:
        # Buscar p贸lizas vencidas o por vencer
        today = datetime.now().date()
        expired_policies = Poliza.objects.filter(
            fecha_fin__lt=today,
            estado__in=['activa', 'suspendida']
        )
        
        renewed_count = 0
        for poliza in expired_policies:
            # Renovar por 1 a帽o
            new_end_date = poliza.fecha_fin + timedelta(days=365)
            poliza.fecha_fin = new_end_date
            poliza.fecha_inicio = poliza.fecha_fin  # Ajustar fecha de inicio
            poliza.estado = 'activa'
            poliza.numero_poliza = f"REN-{poliza.numero_poliza}"
            poliza.save()
            renewed_count += 1
        
        logger.info(f"Se renovaron {renewed_count} p贸lizas")
        
        return JsonResponse({
            'success': True, 
            'message': f'Se renovaron {renewed_count} p贸lizas exitosamente',
            'renewed_count': renewed_count
        })
        
    except Exception as e:
        logger.error(f"Error renovando p贸lizas vencidas: {e}")
        return JsonResponse({'success': False, 'message': f'Error al renovar: {str(e)}'})