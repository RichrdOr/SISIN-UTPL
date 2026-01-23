<<<<<<< Updated upstream
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from datetime import date

from .models import Siniestro
from .forms import SiniestroForm


def crear_siniestro(request):
    if request.method == "POST":
        form = SiniestroForm(request.POST, request.FILES)
        if form.is_valid():
            siniestro = form.save(commit=False)
            siniestro.fecha_apertura = date.today()
            siniestro.save()

            messages.success(request, "Siniestro creado correctamente.")
            return redirect('dashboard_siniestros')
    else:
        form = SiniestroForm()

    return render(request, 'asesora/crear_siniestro.html', {
        'form': form,
        'today': date.today().isoformat(),
    })




def dashboard_siniestros(request):
    siniestros = Siniestro.objects.all().order_by('-fecha_apertura')

    siniestros_data = []
    for s in siniestros:
        siniestros_data.append({
            "id": s.id,
            "tipo_bien": s.tipo_bien,
            "fecha": s.fecha_ocurrencia,
            "estado": s.get_estado_display(),
        })

    stats = {
        "total": siniestros.count(),
        "creados": siniestros.filter(estado='reportado').count(),
        "en_proceso": siniestros.filter(estado='en_revision').count(),
        "finalizados": siniestros.filter(estado__in=['aprobado', 'pagado']).count(),
    }

    return render(request, "siniestros/dashboard.html", {
        "siniestros": siniestros_data,
        "stats": stats,
    })



def detalle_siniestro(request, siniestro_id):
    siniestro = get_object_or_404(Siniestro, id=siniestro_id)

    return render(request, 'siniestros/detalle_siniestro.html', {
        'siniestro': siniestro,
        'documentos': siniestro.documentos.all(),
        'pagare': getattr(siniestro, 'pagare', None)
    })


from django.contrib import messages

def enviar_a_revision(request, siniestro_id):
    siniestro = get_object_or_404(Siniestro, id=siniestro_id)

    try:
        siniestro.revisar()
        siniestro.save()
        messages.info(request, "Siniestro enviado a revisión.")
    except:
        messages.error(request, "No se pudo cambiar el estado.")

    return redirect('detalle_siniestro', siniestro_id=siniestro.id)


def aprobar_siniestro(request, siniestro_id):
    siniestro = get_object_or_404(Siniestro, id=siniestro_id)

    try:
        siniestro.aprobar()
        siniestro.cobertura_valida = True
        siniestro.fecha_respuesta_aseguradora = date.today()
        siniestro.save()
        messages.success(request, "Siniestro aprobado.")
    except:
        messages.error(request, "No se pudo aprobar el siniestro.")

    return redirect('detalle_siniestro', siniestro_id=siniestro.id)


def dashboard_asesora(request):
    return render(request, "asesora/dashboard.html")

def siniestros_asesora(request):
    return render(request, "asesora/siniestros.html")
    
=======
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.db import transaction
from django.db.models import Count, Avg, Sum
from django.utils import timezone
from datetime import date, timedelta
import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# Modelos y Forms
from .models import Siniestro, DocumentoSiniestro, RoboSiniestro, HistorialEstado
from .forms import (
    SiniestroForm, DocumentoSiniestroForm, RoboSiniestroForm,
    DocsIncompletosForm, EnviarAseguradoraForm, AprobarSiniestroForm,
    RechazarSiniestroForm, LiquidarSiniestroForm, RegistrarPagoForm,
    CerrarSiniestroForm
)
from .services import SiniestroService, SiniestroServiceError
from polizas.models import Poliza, RamoPoliza
from usuarios.models import AsesorUTPL
from notificaciones.models import Notificacion
from django.db.models.functions import TruncMonth


# ==========================================
# VISTAS DE DASHBOARD Y LISTADOS (ASESORA)
# ==========================================

def dashboard_siniestros(request):
    """Dashboard principal con métricas, alertas, gráficos y colores corregidos."""
    hoy = date.today()
    fecha_limite = hoy + timedelta(days=30)
    
    # ==========================================
    # 1. MÉTRICAS PRINCIPALES (STATS)
    # ==========================================
    total_siniestros = Siniestro.objects.count()
    siniestros_reportados = Siniestro.objects.filter(estado='reportado').count()
    siniestros_revision = Siniestro.objects.filter(estado='en_revision').count()
    siniestros_aprobados = Siniestro.objects.filter(estado='aprobado').count()
    siniestros_pagados = Siniestro.objects.filter(estado='pagado').count()
    siniestros_rechazados = Siniestro.objects.filter(estado='rechazado').count()
    
    polizas_activas = Poliza.objects.filter(estado='activa').count()
    polizas_por_vencer = Poliza.objects.filter(
        fecha_vencimiento__gte=hoy,
        fecha_vencimiento__lte=fecha_limite
    ).count()
    
    # Tiempo promedio
    siniestros_cerrados = Siniestro.objects.filter(
        estado__in=['aprobado', 'pagado', 'cerrado'],
        tiempo_resolucion_dias__isnull=False
    )
    tiempo_promedio = siniestros_cerrados.aggregate(promedio=Avg('tiempo_resolucion_dias'))['promedio'] or 0
    
    # Tasa aprobación
    total_procesados = siniestros_aprobados + siniestros_pagados + siniestros_rechazados
    tasa_aprobacion = round((siniestros_aprobados + siniestros_pagados) / total_procesados * 100, 1) if total_procesados > 0 else 0
    
    stats = {
        'total': total_siniestros,
        'reportados': siniestros_reportados,
        'en_revision': siniestros_revision,
        'aprobados': siniestros_aprobados,
        'pagados': siniestros_pagados,
        'rechazados': siniestros_rechazados,
        'polizas_activas': polizas_activas,
        'polizas_por_vencer': polizas_por_vencer,
        'tiempo_promedio': round(tiempo_promedio, 1),
        'tasa_aprobacion': tasa_aprobacion,
    }

    # ==========================================
    # 2. CONFIGURACIÓN DE COLORES PARA TABLA Y GRÁFICOS
    # ==========================================
    # Mapeo de (estado_bd -> configuración visual)
    config_estados = {
        'reportado': {'bg': '#f3f4f6', 'color': '#374151', 'label': 'Reportado'}, # Gris
        'docs_incompletos': {'bg': '#ffedd5', 'color': '#c2410c', 'label': 'Docs Incompletos'}, # Naranja
        'documentos_incompletos': {'bg': '#ffedd5', 'color': '#c2410c', 'label': 'Docs Incompletos'}, # Alias
        'docs_completos': {'bg': '#dbeafe', 'color': '#1d4ed8', 'label': 'Docs Completos'}, # Azul claro
        'documentos_completos': {'bg': '#dbeafe', 'color': '#1d4ed8', 'label': 'Docs Completos'}, # Alias
        'enviado': {'bg': '#e0e7ff', 'color': '#4338ca', 'label': 'Enviado'}, # Indigo
        'en_revision': {'bg': '#cffafe', 'color': '#0e7490', 'label': 'En Revisión'}, # Cyan
        'aprobado': {'bg': '#dcfce7', 'color': '#15803d', 'label': 'Aprobado'}, # Verde
        'pagado': {'bg': '#bbf7d0', 'color': '#166534', 'label': 'Pagado'}, # Verde fuerte
        'liquidado': {'bg': '#bae6fd', 'color': '#0369a1', 'label': 'Liquidado'}, # Azul
        'rechazado': {'bg': '#fee2e2', 'color': '#b91c1c', 'label': 'Rechazado'}, # Rojo
        'cerrado': {'bg': '#f3f4f6', 'color': '#6b7280', 'label': 'Cerrado'}, # Gris oscuro
        'fuera_plazo': {'bg': '#fee2e2', 'color': '#991b1b', 'label': 'Fuera de Plazo'}, # Rojo oscuro
    }

    # ==========================================
    # 3. TABLA: ÚLTIMOS SINIESTROS (PROCESADA)
    # ==========================================
    ultimos_siniestros = Siniestro.objects.select_related('poliza').order_by('-fecha_reporte')[:5]

    # Inyectamos los colores dinámicamente
    for s in ultimos_siniestros:
        # Buscamos la config usando el estado actual, si no existe usa 'reportado' por defecto
        config = config_estados.get(s.estado, config_estados['reportado'])
        
        # Asignamos atributos temporales al objeto para que el template los lea
        s.estado_bg = config['bg']
        s.estado_color = config['color']
        s.estado_label = config['label']

    # ==========================================
    # 4. GRÁFICO: EVOLUCIÓN MENSUAL
    # ==========================================
    evolucion_qs = (
        Siniestro.objects
        .annotate(mes=TruncMonth('fecha_reporte'))
        .values('mes')
        .annotate(total=Count('id'))
        .order_by('mes')
    )

    chart_evolucion = {
        'labels': [e['mes'].strftime('%B %Y') if e['mes'] else 'Sin Fecha' for e in evolucion_qs], 
        'data': [e['total'] for e in evolucion_qs]
    }

    # ==========================================
    # 5. GRÁFICO: ESTADOS (DONUT)
    # ==========================================
    estados_qs = Siniestro.objects.values('estado').annotate(total=Count('id'))
    
    chart_estados = {
        'labels': [],
        'data': [],
        'colors': []
    }
    
    for item in estados_qs:
        key = item['estado']
        count = item['total']
        conf = config_estados.get(key, config_estados['reportado'])
        
        chart_estados['labels'].append(conf['label'])
        chart_estados['data'].append(count)
        # Usamos el color de texto ('color') o fondo ('bg') para el gráfico
        chart_estados['colors'].append(conf['color']) 

    # ==========================================
    # 6. GRÁFICO: TIPOS DE EVENTO
    # ==========================================
    tipos_qs = Siniestro.objects.values('tipo_evento').annotate(total=Count('id'))
    chart_tipos = {
        'labels': [t['tipo_evento'].capitalize().replace('_', ' ') for t in tipos_qs],
        'data': [t['total'] for t in tipos_qs],
        'colors': ['#6366f1', '#f97316', '#10b981', '#ef4444', '#3b82f6', '#8b5cf6'] 
    }

    # ==========================================
    # 7. GRÁFICO: MONTOS
    # ==========================================
    montos = Siniestro.objects.aggregate(
        reclamado=Sum('monto_reclamado'),
        aprobado=Sum('monto_aprobado'),
        pagado=Sum('monto_a_pagar')
    )
    chart_montos = {
        'labels': ['Reclamado', 'Aprobado', 'Pagado'],
        'data': [
            float(montos['reclamado'] or 0), 
            float(montos['aprobado'] or 0), 
            float(montos['pagado'] or 0)
        ]
    }

    # ==========================================
    # 8. LISTAS ADICIONALES (Pendientes, Vencer, Notif)
    # ==========================================
    pendientes = Siniestro.objects.filter(
        estado__in=['reportado', 'docs_incompletos', 'docs_completos', 'enviado', 'en_revision']
    ).order_by('fecha_reporte')[:5]
    
    # Calculamos días transcurridos para pendientes
    for p in pendientes:
        p.dias_transcurridos = (date.today() - p.fecha_reporte).days

    proximas_vencer = Poliza.objects.filter(
        fecha_vencimiento__gte=hoy,
        fecha_vencimiento__lte=fecha_limite
    ).order_by('fecha_vencimiento')[:5]
    
    for p in proximas_vencer:
        p.dias_para_vencer = (p.fecha_vencimiento - hoy).days
        # Color de urgencia
        p.urgencia_color = '#ef4444' if p.dias_para_vencer < 7 else '#f59e0b'

    notificaciones = Notificacion.objects.order_by('-fecha_creacion')[:5]
    # Asignar iconos/colores básicos a notificaciones si no tienen
    for n in notificaciones:
        if not hasattr(n, 'color'): n.color = '#3b82f6'
        if not hasattr(n, 'icono'): n.icono = 'bell'

    # Contexto final para el template
    context = {
        'stats': stats,
        'ultimos_siniestros': ultimos_siniestros,
        'chart_evolucion': json.dumps(chart_evolucion),
        'chart_estados': json.dumps(chart_estados),
        'chart_tipos': json.dumps(chart_tipos),
        'chart_montos': json.dumps(chart_montos),
        'pendientes': pendientes,
        'proximas_vencer': proximas_vencer,
        'notificaciones': notificaciones,
    }
    
    return render(request, "asesora/dashboard.html", context)


def dashboard_asesora(request):
    """Dashboard principal para la asesora con métricas, gráficos y listados."""
    hoy = date.today()
    
    # Base querysets
    siniestros = Siniestro.objects.select_related('poliza', 'ramo').all()
    polizas = Poliza.objects.all()

    # Métricas principales
    total = siniestros.count()
    reportados = siniestros.filter(estado='reportado').count()
    docs_incompletos = siniestros.filter(estado='docs_incompletos').count()
    docs_completos = siniestros.filter(estado='docs_completos').count()
    enviados = siniestros.filter(estado='enviado').count()
    en_revision = siniestros.filter(estado='en_revision').count()
    aprobados = siniestros.filter(estado='aprobado').count()
    rechazados = siniestros.filter(estado='rechazado').count()
    liquidados = siniestros.filter(estado='liquidado').count()
    pagados = siniestros.filter(estado='pagado').count()
    cerrados = siniestros.filter(estado='cerrado').count()
    
    # Tiempo promedio de resolución
    siniestros_cerrados = siniestros.filter(
        estado__in=['pagado', 'cerrado'],
        tiempo_resolucion_dias__isnull=False
    )
    tiempo_promedio = siniestros_cerrados.aggregate(
        promedio=Avg('tiempo_resolucion_dias')
    )['promedio'] or 0
    
    # Tasa de aprobación
    total_procesados = aprobados + pagados + rechazados + liquidados + cerrados
    tasa_aprobacion = round(
        (aprobados + pagados + liquidados + cerrados) / total_procesados * 100, 1
    ) if total_procesados > 0 else 0
    
    # Pólizas por vencer (próximos 30 días)
    fecha_limite = hoy + timedelta(days=30)
    polizas_por_vencer = polizas.filter(
        fecha_vencimiento__gte=hoy,
        fecha_vencimiento__lte=fecha_limite
    ).count()

    stats = {
        'total': total,
        'reportados': reportados + docs_incompletos + docs_completos,
        'en_revision': en_revision + enviados,
        'aprobados': aprobados + liquidados,
        'rechazados': rechazados,
        'pagados': pagados + cerrados,
        'polizas_activas': polizas.filter(estado='activa').count(),
        'polizas_por_vencer': polizas_por_vencer,
        'tiempo_promedio': round(tiempo_promedio, 1),
        'tasa_aprobacion': tasa_aprobacion
    }

    # Gráfico 1: Estados (Donut)
    estados = siniestros.values('estado').annotate(total=Count('id'))
    
    estado_labels = {
        'reportado': 'Reportado',
        'docs_incompletos': 'Docs Incompletos',
        'docs_completos': 'Docs Completos',
        'enviado': 'Enviado',
        'en_revision': 'En Revisión',
        'aprobado': 'Aprobado',
        'rechazado': 'Rechazado',
        'liquidado': 'Liquidado',
        'pagado': 'Pagado',
        'cerrado': 'Cerrado',
        'fuera_plazo': 'Fuera de Plazo',
    }
    
    estado_colors = {
        'reportado': '#f59e0b',
        'docs_incompletos': '#ef4444',
        'docs_completos': '#3b82f6',
        'enviado': '#8b5cf6',
        'en_revision': '#06b6d4',
        'aprobado': '#10b981',
        'rechazado': '#dc2626',
        'liquidado': '#0ea5e9',
        'pagado': '#22c55e',
        'cerrado': '#6b7280',
        'fuera_plazo': '#991b1b',
    }

    chart_estados = {
        'labels': [estado_labels.get(e['estado'], e['estado']) for e in estados],
        'data': [e['total'] for e in estados],
        'colors': [estado_colors.get(e['estado'], '#9ca3af') for e in estados]
    }

    # Gráfico 2: Evolución Mensual
    evolucion = (
        siniestros
        .annotate(mes=TruncMonth('fecha_reporte'))
        .values('mes')
        .annotate(total=Count('id'))
        .order_by('mes')
    )

    chart_evolucion = {
        'labels': [e['mes'].strftime('%b %Y') if e['mes'] else 'Sin fecha' for e in evolucion],
        'data': [e['total'] for e in evolucion]
    }

    # Gráfico 3: Tipos de Evento
    tipos = siniestros.values('tipo_evento').annotate(total=Count('id'))
    
    tipo_labels = {
        'danio': 'Daño',
        'robo': 'Robo',
        'hurto': 'Hurto',
        'incendio': 'Incendio',
        'inundacion': 'Inundación',
        'terremoto': 'Terremoto',
        'otro': 'Otro',
    }
    
    tipo_colors = ['#6366f1', '#f97316', '#10b981', '#ef4444', '#3b82f6', '#8b5cf6', '#6b7280']

    chart_tipos = {
        'labels': [tipo_labels.get(t['tipo_evento'], t['tipo_evento']) for t in tipos],
        'data': [t['total'] for t in tipos],
        'colors': tipo_colors[:len(tipos)]
    }

    # Gráfico 4: Montos
    montos = siniestros.aggregate(
        reclamado=Sum('monto_reclamado'),
        aprobado=Sum('monto_aprobado'),
        pagado=Sum('monto_a_pagar')
    )

    chart_montos = {
        'labels': ['Reclamado', 'Aprobado', 'Pagado'],
        'data': [
            float(montos['reclamado'] or 0),
            float(montos['aprobado'] or 0),
            float(montos['pagado'] or 0),
        ],
        'colors': ['#0ea5e9', '#22c55e', '#16a34a']
    }

    # Listados
    ultimos_siniestros = siniestros.order_by('-fecha_reporte')[:5]
    
    pendientes = siniestros.filter(
        estado__in=['reportado', 'docs_incompletos', 'docs_completos', 'enviado', 'en_revision']
    ).order_by('-fecha_reporte')[:5]

    proximas_vencer = polizas.filter(
        fecha_vencimiento__gte=hoy,
        fecha_vencimiento__lte=fecha_limite
    ).order_by('fecha_vencimiento')[:5]

    notificaciones = Notificacion.objects.order_by('-fecha_creacion')[:5]

    context = {
        'stats': stats,
        'chart_estados': json.dumps(chart_estados),
        'chart_evolucion': json.dumps(chart_evolucion),
        'chart_tipos': json.dumps(chart_tipos),
        'chart_montos': json.dumps(chart_montos),
        'ultimos_siniestros': ultimos_siniestros,
        'pendientes': pendientes,
        'proximas_vencer': proximas_vencer,
        'notificaciones': notificaciones,
    }

    return render(request, 'asesora/dashboard.html', context)


def dashboard_asesora(request):
    """Dashboard principal para la asesora con métricas, gráficos y listados."""
    hoy = date.today()
    
    # Base querysets
    siniestros = Siniestro.objects.select_related('poliza', 'ramo').all()
    polizas = Poliza.objects.all()

    # Métricas principales
    total = siniestros.count()
    reportados = siniestros.filter(estado='reportado').count()
    docs_incompletos = siniestros.filter(estado='docs_incompletos').count()
    docs_completos = siniestros.filter(estado='docs_completos').count()
    enviados = siniestros.filter(estado='enviado').count()
    en_revision = siniestros.filter(estado='en_revision').count()
    aprobados = siniestros.filter(estado='aprobado').count()
    rechazados = siniestros.filter(estado='rechazado').count()
    liquidados = siniestros.filter(estado='liquidado').count()
    pagados = siniestros.filter(estado='pagado').count()
    cerrados = siniestros.filter(estado='cerrado').count()
    
    # Tiempo promedio de resolución
    siniestros_cerrados = siniestros.filter(
        estado__in=['pagado', 'cerrado'],
        tiempo_resolucion_dias__isnull=False
    )
    tiempo_promedio = siniestros_cerrados.aggregate(
        promedio=Avg('tiempo_resolucion_dias')
    )['promedio'] or 0
    
    # Tasa de aprobación
    total_procesados = aprobados + pagados + rechazados + liquidados + cerrados
    tasa_aprobacion = round(
        (aprobados + pagados + liquidados + cerrados) / total_procesados * 100, 1
    ) if total_procesados > 0 else 0
    
    # Pólizas por vencer (próximos 30 días)
    fecha_limite = hoy + timedelta(days=30)
    polizas_por_vencer = polizas.filter(
        fecha_vencimiento__gte=hoy,
        fecha_vencimiento__lte=fecha_limite
    ).count()

    stats = {
        'total': total,
        'reportados': reportados + docs_incompletos + docs_completos,
        'en_revision': en_revision + enviados,
        'aprobados': aprobados + liquidados,
        'rechazados': rechazados,
        'pagados': pagados + cerrados,
        'polizas_activas': polizas.filter(estado='activa').count(),
        'polizas_por_vencer': polizas_por_vencer,
        'tiempo_promedio': round(tiempo_promedio, 1),
        'tasa_aprobacion': tasa_aprobacion
    }

    # Gráfico 1: Estados (Donut)
    estados = siniestros.values('estado').annotate(total=Count('id'))
    
    estado_labels = {
        'reportado': 'Reportado',
        'docs_incompletos': 'Docs Incompletos',
        'docs_completos': 'Docs Completos',
        'enviado': 'Enviado',
        'en_revision': 'En Revisión',
        'aprobado': 'Aprobado',
        'rechazado': 'Rechazado',
        'liquidado': 'Liquidado',
        'pagado': 'Pagado',
        'cerrado': 'Cerrado',
        'fuera_plazo': 'Fuera de Plazo',
    }
    
    estado_colors = {
        'reportado': '#f59e0b',
        'docs_incompletos': '#ef4444',
        'docs_completos': '#3b82f6',
        'enviado': '#8b5cf6',
        'en_revision': '#06b6d4',
        'aprobado': '#10b981',
        'rechazado': '#dc2626',
        'liquidado': '#0ea5e9',
        'pagado': '#22c55e',
        'cerrado': '#6b7280',
        'fuera_plazo': '#991b1b',
    }

    chart_estados = {
        'labels': [estado_labels.get(e['estado'], e['estado']) for e in estados],
        'data': [e['total'] for e in estados],
        'colors': [estado_colors.get(e['estado'], '#9ca3af') for e in estados]
    }

    # Gráfico 2: Evolución Mensual
    evolucion = (
        siniestros
        .annotate(mes=TruncMonth('fecha_reporte'))
        .values('mes')
        .annotate(total=Count('id'))
        .order_by('mes')
    )

    chart_evolucion = {
        'labels': [e['mes'].strftime('%b %Y') if e['mes'] else 'Sin fecha' for e in evolucion],
        'data': [e['total'] for e in evolucion]
    }

    # Gráfico 3: Tipos de Evento
    tipos = siniestros.values('tipo_evento').annotate(total=Count('id'))
    
    tipo_labels = {
        'danio': 'Daño',
        'robo': 'Robo',
        'hurto': 'Hurto',
        'incendio': 'Incendio',
        'inundacion': 'Inundación',
        'terremoto': 'Terremoto',
        'otro': 'Otro',
    }
    
    tipo_colors = ['#6366f1', '#f97316', '#10b981', '#ef4444', '#3b82f6', '#8b5cf6', '#6b7280']

    chart_tipos = {
        'labels': [tipo_labels.get(t['tipo_evento'], t['tipo_evento']) for t in tipos],
        'data': [t['total'] for t in tipos],
        'colors': tipo_colors[:len(tipos)]
    }

    # Gráfico 4: Montos
    montos = siniestros.aggregate(
        reclamado=Sum('monto_reclamado'),
        aprobado=Sum('monto_aprobado'),
        pagado=Sum('monto_a_pagar')
    )

    chart_montos = {
        'labels': ['Reclamado', 'Aprobado', 'Pagado'],
        'data': [
            float(montos['reclamado'] or 0),
            float(montos['aprobado'] or 0),
            float(montos['pagado'] or 0),
        ],
        'colors': ['#0ea5e9', '#22c55e', '#16a34a']
    }

    # Listados
    ultimos_siniestros = siniestros.order_by('-fecha_reporte')[:5]
    
    pendientes = siniestros.filter(
        estado__in=['reportado', 'docs_incompletos', 'docs_completos', 'enviado', 'en_revision']
    ).order_by('-fecha_reporte')[:5]

    proximas_vencer = polizas.filter(
        fecha_vencimiento__gte=hoy,
        fecha_vencimiento__lte=fecha_limite
    ).order_by('fecha_vencimiento')[:5]

    notificaciones = Notificacion.objects.order_by('-fecha_creacion')[:5]

    context = {
        'stats': stats,
        'chart_estados': json.dumps(chart_estados),
        'chart_evolucion': json.dumps(chart_evolucion),
        'chart_tipos': json.dumps(chart_tipos),
        'chart_montos': json.dumps(chart_montos),
        'ultimos_siniestros': ultimos_siniestros,
        'pendientes': pendientes,
        'proximas_vencer': proximas_vencer,
        'notificaciones': notificaciones,
    }

    return render(request, 'asesora/dashboard.html', context)


def siniestros_asesora(request):
    """Listado general de siniestros para la asesora"""
    siniestros = Siniestro.objects.select_related('poliza', 'ramo', 'asesor_asignado').order_by('-fecha_reporte')
    
    # Calcular estadísticas
    siniestros_revision_count = siniestros.filter(estado__in=['en_revision', 'enviado']).count()
    siniestros_finalizados_count = siniestros.filter(estado__in=['aprobado', 'pagado', 'cerrado']).count()
    
    context = {
        'siniestros': siniestros,
        'siniestros_revision_count': siniestros_revision_count,
        'siniestros_finalizados_count': siniestros_finalizados_count,
    }
    return render(request, "asesora/siniestros.html", context)


# ==========================================
# CREACIÓN DE SINIESTROS
# ==========================================

def validar_archivo_pdf(archivo):
    """Valida que el archivo sea un PDF válido"""
    if not archivo:
        return True
    
    nombre = archivo.name.lower()
    if not nombre.endswith('.pdf'):
        return False
    
    content_type = archivo.content_type
    if content_type != 'application/pdf':
        return False
    
    return True


def crear_siniestro(request):
    """Vista para crear siniestros y procesar sus documentos"""
    if request.method == "POST":
        form = SiniestroForm(request.POST, request.FILES)
        
        if form.is_valid():
            tipos_documentos = {
                'doc_carta': 'carta',
                'doc_informe': 'informe',
                'doc_proforma': 'proforma',
                'doc_preexistencia': 'preexistencia',
                'doc_denuncia': 'denuncia'
            }
            
            archivos_invalidos = []
            for campo_html, tipo_db in tipos_documentos.items():
                archivo = request.FILES.get(campo_html)
                if archivo and not validar_archivo_pdf(archivo):
                    archivos_invalidos.append(f"{tipo_db.capitalize()}")
            
            if archivos_invalidos:
                messages.error(request, f"Los siguientes documentos no son archivos PDF válidos: {', '.join(archivos_invalidos)}.")
                return render(request, 'asesora/crear_siniestro.html', {
                    'form': form,
                    'today': date.today().isoformat(),
                    'polizas': Poliza.objects.filter(estado='activa'),
                })
            
            try:
                with transaction.atomic():
                    siniestro = form.save(commit=False)
                    siniestro.fecha_apertura = date.today()
                    
                    if not siniestro.ramo_id:
                        siniestro.ramo_id = request.POST.get('ramo')
                    
                    siniestro.save()

                    if request.POST.get('tipo_evento') == 'robo':
                        RoboSiniestro.objects.create(
                            siniestro=siniestro,
                            denuncia_policial=request.POST.get('denuncia_policial'),
                            fiscalia=request.POST.get('fiscalia'),
                            fecha_denuncia=request.POST.get('fecha_denuncia') or date.today()
                        )

                    for campo_html, tipo_db in tipos_documentos.items():
                        archivo = request.FILES.get(campo_html)
                        if archivo:
                            DocumentoSiniestro.objects.create(
                                siniestro=siniestro,
                                tipo=tipo_db,
                                archivo=archivo
                            )
                    
                    # Registrar en historial
                    HistorialEstado.objects.create(
                        siniestro=siniestro,
                        estado_anterior='',
                        estado_nuevo='reportado',
                        usuario=request.user.username if request.user.is_authenticated else 'Sistema',
                        observaciones='Siniestro creado'
                    )
                    
                    messages.success(request, f"¡Siniestro {siniestro.numero_siniestro} creado exitosamente!")
                    return redirect('siniestros:detalle_siniestro', siniestro_id=siniestro.id)
                    
            except Exception as e:
                messages.error(request, f"Error al guardar: {e}")
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"Error en {field}: {error}")
    else:
        form = SiniestroForm()

    return render(request, 'asesora/crear_siniestro.html', {
        'form': form,
        'today': date.today().isoformat(),
        'polizas': Poliza.objects.filter(estado='activa'),
    })


# ==========================================
# DETALLE Y GESTIÓN DE SINIESTROS
# ==========================================

def detalle_siniestro(request, siniestro_id):
    """Vista de detalle del siniestro con acciones disponibles"""
    siniestro = get_object_or_404(Siniestro, id=siniestro_id)
    
    # Definir los estados del timeline según el nuevo flujo
    estados_orden = [
        'reportado', 'docs_incompletos', 'docs_completos', 
        'enviado', 'en_revision', 'aprobado', 'liquidado', 'pagado', 'cerrado'
    ]
    estado_actual = siniestro.estado
    
    # Calcular el índice del estado actual
    try:
        indice_actual = estados_orden.index(estado_actual)
    except ValueError:
        indice_actual = 0
    
    # Calcular progreso
    progreso_estado = (indice_actual / (len(estados_orden) - 1)) * 100 if len(estados_orden) > 1 else 0
    
    # Construir lista de estados para el timeline
    estados = [
        {'nombre': 'Reportado', 'icono': '📝', 'key': 'reportado'},
        {'nombre': 'Docs', 'icono': '📋', 'key': 'docs_completos'},
        {'nombre': 'Enviado', 'icono': '📤', 'key': 'enviado'},
        {'nombre': 'Revisión', 'icono': '🔍', 'key': 'en_revision'},
        {'nombre': 'Aprobado', 'icono': '✅', 'key': 'aprobado'},
        {'nombre': 'Liquidado', 'icono': '💰', 'key': 'liquidado'},
        {'nombre': 'Pagado', 'icono': '💳', 'key': 'pagado'},
        {'nombre': 'Cerrado', 'icono': '🔒', 'key': 'cerrado'},
    ]
    
    for estado in estados:
        try:
            estado_idx = estados_orden.index(estado['key'])
            estado['activo'] = indice_actual >= estado_idx
            estado['actual'] = estado_actual == estado['key']
        except ValueError:
            estado['activo'] = False
            estado['actual'] = False
    
    # Si está rechazado, mostrar timeline especial
    if estado_actual == 'rechazado':
        estados = [
            {'nombre': 'Reportado', 'icono': '📝', 'activo': True, 'actual': False},
            {'nombre': 'Revisión', 'icono': '🔍', 'activo': True, 'actual': False},
            {'nombre': 'Rechazado', 'icono': '❌', 'activo': True, 'actual': True},
            {'nombre': '-', 'icono': '⬜', 'activo': False, 'actual': False},
        ]
        progreso_estado = 33
    
    # Obtener acciones disponibles
    acciones = SiniestroService.obtener_acciones_disponibles(siniestro)
    
    # Formularios para los modals
    form_documento = DocumentoSiniestroForm()
    form_docs_incompletos = DocsIncompletosForm()
    form_enviar_aseguradora = EnviarAseguradoraForm(initial={
        'correo_aseguradora': siniestro.poliza.aseguradora if siniestro.poliza else ''
    })
    form_aprobar = AprobarSiniestroForm()
    form_rechazar = RechazarSiniestroForm()
    form_liquidar = LiquidarSiniestroForm()
    form_pago = RegistrarPagoForm()
    form_cerrar = CerrarSiniestroForm()
    
    # Historial de estados
    historial = siniestro.historial_estados.all()[:10]
    
    return render(request, 'asesora/detalle_siniestro.html', {
        'siniestro': siniestro,
        'documentos': siniestro.documentos.all(),
        'estados': estados,
        'progreso_estado': progreso_estado,
        'acciones': acciones,
        'historial': historial,
        # Formularios
        'form_documento': form_documento,
        'form_docs_incompletos': form_docs_incompletos,
        'form_enviar_aseguradora': form_enviar_aseguradora,
        'form_aprobar': form_aprobar,
        'form_rechazar': form_rechazar,
        'form_liquidar': form_liquidar,
        'form_pago': form_pago,
        'form_cerrar': form_cerrar,
    })


# ==========================================
# SUBIDA DE DOCUMENTOS
# ==========================================

@require_http_methods(["POST"])
def subir_documento(request, siniestro_id):
    """Vista para subir documentos a un siniestro"""
    siniestro = get_object_or_404(Siniestro, id=siniestro_id)
    
    if siniestro.estado == 'cerrado':
        messages.error(request, "No se pueden subir documentos a un siniestro cerrado.")
        return redirect('siniestros:detalle_siniestro', siniestro_id=siniestro.id)
    
    form = DocumentoSiniestroForm(request.POST, request.FILES)
    
    if form.is_valid():
        archivo = form.cleaned_data['archivo']
        
        # Validar que sea PDF
        if not validar_archivo_pdf(archivo):
            messages.error(request, "Solo se permiten archivos PDF.")
            return redirect('siniestros:detalle_siniestro', siniestro_id=siniestro.id)
        
        try:
            documento = form.save(commit=False)
            documento.siniestro = siniestro
            documento.save()
            
            messages.success(request, f"Documento '{documento.get_tipo_display()}' subido exitosamente.")
        except Exception as e:
            messages.error(request, f"Error al subir documento: {e}")
    else:
        for field, errors in form.errors.items():
            for error in errors:
                messages.error(request, f"Error en {field}: {error}")
    
    return redirect('siniestros:detalle_siniestro', siniestro_id=siniestro.id)


# ==========================================
# TRANSICIONES DE ESTADO (usando SiniestroService)
# ==========================================

@require_http_methods(["POST"])
def marcar_docs_incompletos(request, siniestro_id):
    """Marca el siniestro como documentos incompletos"""
    siniestro = get_object_or_404(Siniestro, id=siniestro_id)
    
    try:
        docs_faltantes = request.POST.getlist('documentos_faltantes')
        observaciones = request.POST.get('observaciones', '')
        
        docs_nombres = {
            'cedula': 'Cédula de identidad',
            'poliza': 'Copia de la póliza',
            'denuncia': 'Denuncia policial',
            'facturas': 'Facturas o comprobantes',
            'fotos': 'Fotografías del siniestro',
            'certificado_medico': 'Certificado médico',
            'informe_tecnico': 'Informe técnico',
            'proforma': 'Proforma de reparación',
            'otros': 'Otros documentos'
        }
        
        docs_texto = [docs_nombres.get(d, d) for d in docs_faltantes]
        usuario = request.user.username if request.user.is_authenticated else 'Sistema'
        
        SiniestroService.marcar_documentos_incompletos(
            siniestro, docs_texto, usuario
        )
        
        messages.warning(request, f"Siniestro marcado como 'Documentos Incompletos'. Se notificará al reclamante.")
        
    except SiniestroServiceError as e:
        messages.error(request, str(e))
    except Exception as e:
        messages.error(request, f"Error: {str(e)}")
    
    return redirect('siniestros:detalle_siniestro', siniestro_id=siniestro.id)


@require_http_methods(["POST"])
def confirmar_documentos(request, siniestro_id):
    """Confirma que los documentos están completos"""
    siniestro = get_object_or_404(Siniestro, id=siniestro_id)
    
    try:
        usuario = request.user.username if request.user.is_authenticated else 'Sistema'
        
        SiniestroService.marcar_documentos_completos(siniestro, usuario)
        
        messages.success(request, "Documentos marcados como completos. El siniestro está listo para enviar a la aseguradora.")
        
    except SiniestroServiceError as e:
        messages.error(request, str(e))
    except Exception as e:
        messages.error(request, f"Error: {str(e)}")
    
    return redirect('siniestros:detalle_siniestro', siniestro_id=siniestro.id)


@require_http_methods(["POST"])
def enviar_recordatorio(request, siniestro_id):
    """Envía recordatorio de documentos pendientes"""
    siniestro = get_object_or_404(Siniestro, id=siniestro_id)
    
    try:
        SiniestroService.enviar_recordatorio_documentos(siniestro)
        messages.info(request, "Recordatorio enviado al reclamante.")
        
    except SiniestroServiceError as e:
        messages.error(request, str(e))
    except Exception as e:
        messages.error(request, f"Error: {str(e)}")
    
    return redirect('siniestros:detalle_siniestro', siniestro_id=siniestro.id)


@require_http_methods(["POST"])
def enviar_aseguradora(request, siniestro_id):
    """Envía el siniestro a la aseguradora"""
    siniestro = get_object_or_404(Siniestro, id=siniestro_id)
    
    try:
        correo = request.POST.get('correo_aseguradora', '')
        mensaje = request.POST.get('mensaje', '')
        usuario = request.user.username if request.user.is_authenticated else 'Sistema'

        if not correo:
            messages.error(request, "Debe especificar el correo de la aseguradora.")
            return redirect('siniestros:detalle_siniestro', siniestro_id=siniestro.id)
        
        SiniestroService.enviar_a_aseguradora(siniestro, correo, mensaje, usuario)
        
        messages.success(request, f"Siniestro enviado a la aseguradora. Fecha límite de respuesta: {siniestro.fecha_limite_respuesta_aseguradora}")
        
    except SiniestroServiceError as e:
        messages.error(request, str(e))
    except Exception as e:
        messages.error(request, f"Error: {str(e)}")
    
    return redirect('siniestros:detalle_siniestro', siniestro_id=siniestro.id)


@require_http_methods(["POST"])
def marcar_revision(request, siniestro_id):
    """Marca el siniestro como en revisión"""
    siniestro = get_object_or_404(Siniestro, id=siniestro_id)
    
    try:
        usuario = request.user.username if request.user.is_authenticated else 'Sistema'
        
        SiniestroService.marcar_en_revision(siniestro, usuario)
        
        messages.info(request, "Siniestro marcado como 'En Revisión' por la aseguradora.")
        
    except SiniestroServiceError as e:
        messages.error(request, str(e))
    except Exception as e:
        messages.error(request, f"Error: {str(e)}")
    
    return redirect('siniestros:detalle_siniestro', siniestro_id=siniestro.id)


@require_http_methods(["POST"])
def aprobar_siniestro_modal(request, siniestro_id):
    """Aprueba el siniestro (solo confirma cobertura, NO montos)"""
    siniestro = get_object_or_404(Siniestro, id=siniestro_id)
    
    try:
        observaciones = request.POST.get('observaciones', '')
        usuario = request.user.username if request.user.is_authenticated else 'Sistema'
        
        SiniestroService.aprobar(siniestro, observaciones, usuario)
        
        messages.success(request, "Siniestro aprobado. Ahora puede proceder con la liquidación.")
        
    except SiniestroServiceError as e:
        messages.error(request, str(e))
    except Exception as e:
        messages.error(request, f"Error: {str(e)}")
    
    return redirect('siniestros:detalle_siniestro', siniestro_id=siniestro.id)


@require_http_methods(["POST"])
def rechazar_siniestro(request, siniestro_id):
    """Rechaza el siniestro"""
    siniestro = get_object_or_404(Siniestro, id=siniestro_id)
    
    try:
        razon_principal = request.POST.get('razon_principal', '')
        detalle = request.POST.get('detalle_rechazo', '')
        usuario = request.user.username if request.user.is_authenticated else 'Sistema'
        
        razones_nombres = {
            'no_cubierto': 'Evento no cubierto por la póliza',
            'fuera_plazo': 'Reportado fuera del plazo establecido',
            'bien_no_coincide': 'El bien no coincide con la póliza',
            'exclusion': 'Aplica exclusión de la póliza',
            'documentacion_invalida': 'Documentación inválida o insuficiente',
            'fraude': 'Sospecha de fraude',
            'otro': 'Otra razón',
        }
        
        razon_texto = razones_nombres.get(razon_principal, razon_principal)
        razon_completa = f"{razon_texto}: {detalle}" if detalle else razon_texto
        
        SiniestroService.rechazar(siniestro, razon_completa, usuario)
        
        messages.error(request, f"Siniestro rechazado. Razón: {razon_texto}")
        
    except SiniestroServiceError as e:
        messages.error(request, str(e))
    except Exception as e:
        messages.error(request, f"Error: {str(e)}")
    
    return redirect('siniestros:detalle_siniestro', siniestro_id=siniestro.id)


@require_http_methods(["POST"])
def liquidar_siniestro(request, siniestro_id):
    """Liquida el siniestro (ingresa montos)"""
    siniestro = get_object_or_404(Siniestro, id=siniestro_id)
    
    try:
        monto_aprobado = request.POST.get('monto_aprobado', 0)
        deducible = request.POST.get('deducible', 0)
        notas = request.POST.get('notas', '')
        documento = request.FILES.get('documento_liquidacion')
        usuario = request.user.username if request.user.is_authenticated else 'Sistema'
        
        # Validar documento si existe
        if documento and not validar_archivo_pdf(documento):
            messages.error(request, "El documento de liquidación debe ser un archivo PDF.")
            return redirect('siniestros:detalle_siniestro', siniestro_id=siniestro.id)
        
        SiniestroService.liquidar(siniestro, monto_aprobado, deducible, notas, usuario)
        
        # Guardar documento de liquidación
        if documento:
            DocumentoSiniestro.objects.create(
                siniestro=siniestro,
                tipo='liquidacion',
                archivo=documento,
                descripcion='Documento de liquidación de aseguradora'
            )
        
        messages.success(request, f"Siniestro liquidado. Monto a pagar: ${siniestro.monto_a_pagar}")
        
    except SiniestroServiceError as e:
        messages.error(request, str(e))
    except Exception as e:
        messages.error(request, f"Error: {str(e)}")
    
    return redirect('siniestros:detalle_siniestro', siniestro_id=siniestro.id)


@require_http_methods(["POST"])
def registrar_pago(request, siniestro_id):
    """Registra el pago del siniestro"""
    siniestro = get_object_or_404(Siniestro, id=siniestro_id)
    
    try:
        comprobante = request.FILES.get('comprobante_pago')
        observaciones = request.POST.get('observaciones', '')
        usuario = request.user.username if request.user.is_authenticated else 'Sistema'
        
        # Validar comprobante
        if comprobante:
            nombre = comprobante.name.lower()
            if not (nombre.endswith('.pdf') or nombre.endswith('.jpg') or 
                    nombre.endswith('.jpeg') or nombre.endswith('.png')):
                messages.error(request, "El comprobante debe ser PDF o imagen (JPG, PNG).")
                return redirect('siniestros:detalle_siniestro', siniestro_id=siniestro.id)
        
        SiniestroService.registrar_pago(siniestro, usuario)
        
        # Guardar comprobante de pago
        if comprobante:
            DocumentoSiniestro.objects.create(
                siniestro=siniestro,
                tipo='comprobante_pago',
                archivo=comprobante,
                descripcion='Comprobante de pago al reclamante'
            )
        
        if observaciones:
            siniestro.observaciones_internas += f"\n[{timezone.now().strftime('%d/%m/%Y %H:%M')}] Pago: {observaciones}"
            siniestro.save()
        
        messages.success(request, "Pago registrado exitosamente. El siniestro puede ser cerrado.")
        
    except SiniestroServiceError as e:
        messages.error(request, str(e))
    except Exception as e:
        messages.error(request, f"Error: {str(e)}")
    
    return redirect('siniestros:detalle_siniestro', siniestro_id=siniestro.id)


@require_http_methods(["POST"])
def cerrar_siniestro(request, siniestro_id):
    """Cierra el siniestro"""
    siniestro = get_object_or_404(Siniestro, id=siniestro_id)
    
    try:
        notas_cierre = request.POST.get('notas_cierre', '')
        usuario = request.user.username if request.user.is_authenticated else 'Sistema'
        
        SiniestroService.cerrar(siniestro, notas_cierre, usuario)
        
        messages.success(request, f"Siniestro {siniestro.numero_siniestro} cerrado exitosamente.")
        
    except SiniestroServiceError as e:
        messages.error(request, str(e))
    except Exception as e:
        messages.error(request, f"Error: {str(e)}")
    
    return redirect('siniestros:detalle_siniestro', siniestro_id=siniestro.id)


# ==========================================
# VISTAS AUXILIARES
# ==========================================

@require_http_methods(["GET"])
def obtener_ramos_poliza(request, poliza_id):
    """API para obtener ramos de una póliza"""
    ramos = RamoPoliza.objects.filter(poliza_id=poliza_id).values('id', 'ramo', 'suma_asegurada')
    return JsonResponse({'success': True, 'ramos': list(ramos)})


def exportar_siniestros_excel(request):
    """Exporta siniestros a Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Siniestros"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    alignment = Alignment(horizontal="center", vertical="center")

    headers = [
        'Nº Siniestro', 'Póliza', 'Titular', 'Aseguradora', 
        'Tipo Evento', 'Fecha Ocurrencia', 'Monto Reclamado', 'Estado'
    ]
    
    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=column_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20

    siniestros = Siniestro.objects.select_related('poliza').all().order_by('-fecha_reporte')

    for row_num, siniestro in enumerate(siniestros, 2):
        ws.cell(row=row_num, column=1, value=siniestro.numero_siniestro)
        ws.cell(row=row_num, column=2, value=siniestro.poliza.numero_poliza if siniestro.poliza else "N/A")
        ws.cell(row=row_num, column=3, value=str(siniestro.poliza.titular) if siniestro.poliza else "N/A")
        ws.cell(row=row_num, column=4, value=siniestro.poliza.aseguradora if siniestro.poliza else "N/A")
        ws.cell(row=row_num, column=5, value=siniestro.get_tipo_evento_display())
        ws.cell(row=row_num, column=6, value=siniestro.fecha_ocurrencia.strftime('%d/%m/%Y') if siniestro.fecha_ocurrencia else "N/A")
        ws.cell(row=row_num, column=7, value=float(siniestro.monto_reclamado or 0))
        ws.cell(row=row_num, column=8, value=siniestro.get_estado_display())

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="Reporte_Siniestros_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    
    wb.save(response)
    return response


# Vistas legacy para compatibilidad
def enviar_a_revision(request, siniestro_id):
    return marcar_revision(request, siniestro_id)

def aprobar_siniestro(request, siniestro_id):
    return aprobar_siniestro_modal(request, siniestro_id)
>>>>>>> Stashed changes
